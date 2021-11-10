from openpyxl import load_workbook
wb2 = load_workbook('c:\\VBA_kurs\\bok1.xlsx')
print(wb2.sheetnames)

sheet  = wb2['Statistik']
for raden in range(2,10):
    namn = sheet.cell(row=raden, column=1).value
    belopp = sheet.cell(row=raden, column=2).value
    bla = sheet.cell(row=raden, column=3).value
    if namn == None:
        break
    print(f"{namn} {belopp} {bla}")
